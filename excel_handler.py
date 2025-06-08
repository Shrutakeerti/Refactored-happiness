import openpyxl
# provides func to extract labels
def extract_labels_from_excel(template_path): # this func extracts labels from the template
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    labels = [] # list to store whatever I am extracting
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().startswith("◦"):
                labels.append(cell.value.strip())
    return labels

def update_excel_with_values(template_path, output_path, value_map):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active # this func returns the active worksheet

    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            label = cell.value.strip() if isinstance(cell.value, str) else None
            if label in value_map:
                # Update first numeric cell with 0 value to mapped value
                for j in range(i+1, len(row)):
                    if isinstance(row[j].value, (int, float)) and row[j].value == 0:
                        row[j].value = value_map[label]
                        break
    wb.save(output_path)
